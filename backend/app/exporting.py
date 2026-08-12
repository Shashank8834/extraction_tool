"""Admin exports: submissions as Excel, uploaded documents as a ZIP.

The per-submission workbook is the office's master file for an incorporation.
Three sheets:

  Details        — what the client submitted, Particular / Details, section by
                   section, mirroring the intake spreadsheet the office already
                   works from.
  LLP agreement  — the drafting inputs, pulled from Details by formula.
  ODI sheet      — the ODI inputs; the Indian side pulled from Details, the
                   foreign-entity and bank side left for the team to fill.

The two derived sheets reference `Details` with real formulas rather than
copied values, so the workbook stays a live master: correct an address on
Details and both derived sheets follow. Row numbers are resolved at write time
from the field names, so adding a field or a third partner to
form_config.yaml cannot leave a formula pointing at the wrong row.
"""
import io
import zipfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .security import sanitize_filename
from .storage import load_decrypted

_FONT = "Arial"
_HEAD = Font(name=_FONT, size=11, bold=True, color="FFFFFF")
_SECTION = Font(name=_FONT, size=11, bold=True)
_LABEL = Font(name=_FONT, size=10, bold=True)
_BODY = Font(name=_FONT, size=10)
_NOTE = Font(name=_FONT, size=9, italic=True, color="666666")
# green: pulled from another sheet by formula — do not type over it
_LINKED = Font(name=_FONT, size=10, color="008000")
_HEAD_FILL = PatternFill("solid", fgColor="6958C2")  # BCL purple
_SECTION_FILL = PatternFill("solid", fgColor="EDEDED")
# yellow: the team fills this in by hand
_INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)

# party naming in the agreement, in the order partners appear in the config
_ORDINALS = ["First", "Second", "Third", "Fourth", "Fifth", "Sixth", "Seventh"]


def _ordinal(i: int) -> str:
    return _ORDINALS[i] if i < len(_ORDINALS) else f"Partner {i + 1}"


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row: int, cells: list, font: Font, fill: PatternFill | None = None) -> None:
    for col, value in enumerate(cells, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.alignment = _WRAP
        cell.border = _BORDER
        if fill is not None:
            cell.fill = fill


def _as_number(field: dict, value):
    """Write a `numeric: true` field as a number so the derived sheets can add
    it up — a text cell counts as zero inside SUM, which silently produces a
    total of 0 and blank percentages.

    Only fields the config marks: coercing everything would strip the leading
    zero off a phone number. Anything that will not parse stays as typed, so a
    client writing "one lakh" is left alone rather than lost.
    """
    if not field.get("numeric") or not isinstance(value, str):
        return value
    cleaned = value.strip().replace(",", "").replace("₹", "").replace("%", "").strip()
    if not cleaned:
        return value
    try:
        number = float(cleaned)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def _details_sheet(ws, cfg: dict, link, submission, uploads_by_slot: dict) -> dict[str, int]:
    """Write the submitted data section by section.

    Returns {field_name: row}, which the derived sheets use to build their
    formulas — so no row number is ever written by hand.
    """
    _write_row(ws, 1, ["Particular", "Details"], _HEAD, _HEAD_FILL)
    ws.freeze_panes = "A2"

    rows: dict[str, int] = {}
    row = 2
    meta = [
        ("Client / reference", link.label or "—"),
        ("Submitted (UTC)", submission.submitted_at.strftime("%Y-%m-%d %H:%M")),
        ("Submitted from IP", submission.client_ip or "—"),
    ]
    for label, value in meta:
        _write_row(ws, row, [label, value], _BODY)
        ws.cell(row=row, column=1).font = _LABEL
        row += 1
    row += 1

    for section in cfg["sections"]:
        _write_row(ws, row, [section["title"], ""], _SECTION, _SECTION_FILL)
        row += 1

        for up in section["uploads"]:
            upload = uploads_by_slot.get(up["name"])
            _write_row(
                ws, row,
                [up["label"], upload.original_filename if upload else "Not uploaded"],
                _BODY,
            )
            ws.cell(row=row, column=1).font = _LABEL
            row += 1

        for field in section["fields"]:
            value = (submission.form_data or {}).get(field["name"]) or ""
            _write_row(ws, row, [field["label"], _as_number(field, value)], _BODY)
            ws.cell(row=row, column=1).font = _LABEL
            rows[field["name"]] = row
            row += 1

        row += 1

    _autosize(ws, {1: 42, 2: 62})
    return rows


def _link_cell(ws, row: int, col: int, ref: str | None) -> None:
    """A cell that pulls its value from Details, in green so nobody types over
    it. A field the config does not have yields a blank rather than a #REF!.

    The IF guard matters: a plain `=Details!B20` pointing at an empty cell
    renders as `0`, not as blank, and a `0` is exactly what must not turn up
    where a father's name or a DIN belongs.
    """
    formula = f'=IF(Details!B{ref}="","",Details!B{ref})' if ref else ""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = _LINKED if ref else _BODY
    cell.alignment = _WRAP
    cell.border = _BORDER


def _input_cell(ws, row: int, col: int) -> None:
    """A blank the team fills in by hand — yellow, per the legend."""
    cell = ws.cell(row=row, column=col, value="")
    cell.font = _BODY
    cell.alignment = _WRAP
    cell.border = _BORDER
    cell.fill = _INPUT_FILL


def _llp_agreement_sheet(ws, cfg: dict, rows: dict[str, int]) -> None:
    """Drafting inputs for the LLP agreement, pulled from Details."""
    partners = [s for s in cfg["sections"] if s["is_partner"]]

    _write_row(ws, 1, ["Particular", "Details", ""], _HEAD, _HEAD_FILL)

    r = 2
    _write_row(ws, r, ["Name of the LLP", "", ""], _BODY)
    ws.cell(row=r, column=1).font = _LABEL
    _link_cell(ws, r, 2, rows.get("llp__proposed_name_1"))
    r += 1
    _write_row(ws, r, ["Registered office address", "", ""], _BODY)
    ws.cell(row=r, column=1).font = _LABEL
    _link_cell(ws, r, 2, rows.get("office__office_address"))
    r += 2

    # one block per partner: name, father's name, address
    for i, section in enumerate(partners):
        key = section["key"]
        _write_row(ws, r, [f"{_ordinal(i)} Party details", "", ""], _SECTION, _SECTION_FILL)
        r += 1
        for label, first, last in (
            ("Name", f"{key}__first_name", f"{key}__last_name"),
            ("Father's name", f"{key}__father_first_name", f"{key}__father_last_name"),
        ):
            _write_row(ws, r, [label, "", ""], _BODY)
            ws.cell(row=r, column=1).font = _LABEL
            _link_cell(ws, r, 2, rows.get(first))
            _link_cell(ws, r, 3, rows.get(last))
            r += 1
        _write_row(ws, r, ["Address", "", ""], _BODY)
        ws.cell(row=r, column=1).font = _LABEL
        _link_cell(ws, r, 2, rows.get(f"{key}__present_address"))
        r += 2

    # capital: each partner's contribution, the total, and each one's share of
    # it as a percentage
    _write_row(ws, r, ["Capital sharing ratio", "Capital", "%"], _SECTION, _SECTION_FILL)
    r += 1
    first_capital_row = r
    for i, _ in enumerate(partners):
        _write_row(ws, r, [f"{_ordinal(i)} Party", "", ""], _BODY)
        ws.cell(row=r, column=1).font = _LABEL
        _link_cell(ws, r, 2, rows.get(f"capital__partner{i + 1}_capital"))
        r += 1
    last_capital_row = r - 1
    total_row = r
    _write_row(ws, r, ["Total", "", ""], _BODY)
    ws.cell(row=r, column=1).font = _LABEL
    ws.cell(row=r, column=2).value = f"=SUM(B{first_capital_row}:B{last_capital_row})"
    ws.cell(row=r, column=2).font = _LABEL
    # percentages only once the total is known, and never a divide-by-zero
    for i in range(len(partners)):
        cap_row = first_capital_row + i
        pct = ws.cell(row=cap_row, column=3)
        pct.value = f'=IFERROR(B{cap_row}/$B${total_row}*100,"")'
        pct.font = _BODY
        pct.number_format = "0.00"
    r += 2

    _write_row(ws, r, ["Profit sharing ratio", "%", ""], _SECTION, _SECTION_FILL)
    r += 1
    for i, _ in enumerate(partners):
        _write_row(ws, r, [f"{_ordinal(i)} Party", "", ""], _BODY)
        ws.cell(row=r, column=1).font = _LABEL
        _link_cell(ws, r, 2, rows.get(f"capital__partner{i + 1}_profit"))
        r += 1

    _autosize(ws, {1: 34, 2: 46, 3: 30})


def _odi_sheet(ws, cfg: dict, rows: dict[str, int]) -> None:
    """ODI inputs. The Indian side comes from Details; the foreign entity and
    the bank details are left yellow for the team."""
    partners = [s for s in cfg["sections"] if s["is_partner"]]
    signatory = partners[0]["key"] if partners else ""

    _write_row(ws, 1, ["Particular", "Details", ""], _HEAD, _HEAD_FILL)
    _write_row(
        ws, 2,
        ["Yellow cells are filled in by the team. Green values come from the "
         "Details sheet — correct them there, not here.", "", ""],
        _NOTE,
    )

    r = 3
    for label, ref in (
        ("Name of the LLP", rows.get("llp__proposed_name_1")),
        ("Registered office address", rows.get("office__office_address")),
        ("Contact number", rows.get("llp_details__llp_contact")),
        ("Email ID of the LLP", rows.get("llp_details__llp_email")),
    ):
        _write_row(ws, r, [label, "", ""], _BODY)
        ws.cell(row=r, column=1).font = _LABEL
        _link_cell(ws, r, 2, ref)
        r += 1
    r += 1

    _write_row(
        ws, r,
        ["Partner authorised to sign the ODI", "", ""], _SECTION, _SECTION_FILL,
    )
    r += 1
    _write_row(ws, r, ["Name", "", ""], _BODY)
    ws.cell(row=r, column=1).font = _LABEL
    _link_cell(ws, r, 2, rows.get(f"{signatory}__first_name"))
    _link_cell(ws, r, 3, rows.get(f"{signatory}__last_name"))
    r += 1
    for label, ref in (
        ("Address", rows.get(f"{signatory}__present_address")),
        ("Email ID", rows.get(f"{signatory}__email")),
        ("Contact number", rows.get(f"{signatory}__mobile")),
    ):
        _write_row(ws, r, [label, "", ""], _BODY)
        ws.cell(row=r, column=1).font = _LABEL
        _link_cell(ws, r, 2, ref)
        r += 1
    r += 1

    # everything below is the team's to fill: nothing in the intake form can
    # supply it
    manual_blocks = [
        ("Foreign entity details (filled in by the team)",
         ["Foreign entity name", "Address", "Email ID", "Contact number"]),
        ("Foreign entity shareholding pattern",
         ["Shareholder 1", "Shareholder 2"]),
        ("Foreign entity bank account details",
         ["Account number", "SWIFT", "IBAN"]),
        ("Indian LLP bank account",
         ["Account number", "IFSC", "Name of the bank"]),
    ]
    for title, labels in manual_blocks:
        _write_row(ws, r, [title, "", ""], _SECTION, _SECTION_FILL)
        r += 1
        for label in labels:
            _write_row(ws, r, [label, "", ""], _BODY)
            ws.cell(row=r, column=1).font = _LABEL
            _input_cell(ws, r, 2)
            r += 1
        r += 1

    _autosize(ws, {1: 44, 2: 46, 3: 30})


# The layout llp-gen's ExcelParser expects. It finds each partner block by
# scanning column A for "designated partner details" and then reads fixed
# offsets from that row, and it reads the LLP-level rows by absolute number —
# so this sheet must keep exactly these labels at exactly these rows.
_LLP_GEN_PARTNER_LABELS = [
    ("First Name", "first_name"),
    ("Last Name", "last_name"),
    ("DIN if any", "din"),
    ("Father First Name", "father_first_name"),
    ("Father Last name", "father_last_name"),
    ("Present address of the Partner", "present_address"),
    ("Permanent Address of the partner", "permanent_address"),
    ("Mobile number", "mobile"),
    ("Email id", "email"),
    ("Occupation", "occupation"),
    ("Income tax Pan/Passport", "pan_number"),
    ("Date of Birth", "dob"),
    ("Nationality", "nationality"),
    ("Director/Partner in any company/LLP", "partner_elsewhere"),
    ("Member of ICSI/ICAI/ICWAI", "professional_membership"),
    ("Relation (Son/Daughter/Wife/Husband)", "relation"),
]


def _llp_gen_sheet(ws, cfg: dict, submission) -> None:
    """The sheet the LLP document generator reads.

    Values, never formulas: llp-gen opens the file with plain openpyxl, which
    hands back a formula as its literal text rather than its result. A formula
    here would put "=Details!B17" into a filed document.
    """
    data = submission.form_data or {}
    partners = [s for s in cfg["sections"] if s["is_partner"]][:2]

    def put(row: int, label: str, value="") -> None:
        ws.cell(row=row, column=1, value=label).font = _LABEL
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _BODY
        cell.alignment = _WRAP

    put(1, "Name of the Proposed LLP", data.get("llp__proposed_name_1", ""))

    # partner blocks at rows 2 and 19, exactly as the generator expects
    for i, section in enumerate(partners):
        head = 2 + i * 17
        ordinal = "First" if i == 0 else "Second"
        put(head, f"{ordinal} Designated partner details")
        for offset, (label, key) in enumerate(_LLP_GEN_PARTNER_LABELS, start=1):
            put(head + offset, label, data.get(f"{section['key']}__{key}", ""))

    put(36, "Capital of the LLP")
    put(37, "First Designated partner share", data.get("capital__partner1_capital", ""))
    put(38, "Second Designated partner share", data.get("capital__partner2_capital", ""))
    put(40, "Profit or loss sharing ratio")
    put(41, "First Designated partner share", data.get("capital__partner1_profit", ""))
    put(42, "Second Designated partner share", data.get("capital__partner2_profit", ""))
    put(44, "Registered office address", data.get("office__office_address", ""))
    put(45, "Property in the name of/name in the Utility bill",
        data.get("office__utility_name", ""))
    put(47, "DSC status - person with a DIN needs a DSC mandatorily, else a DSC of "
            "one partner will do", data.get("partner1__has_dsc", ""))
    put(48, "DSC Registered in MCA?")
    put(49, "Signed documents in place")
    put(51, "Objects of the LLP", data.get("llp_details__object", ""))
    put(53, "Email id and contact number of the LLP, other than one given above",
        f'{data.get("llp_details__llp_email", "")} / '
        f'{data.get("llp_details__llp_contact", "")}'.strip(" /"))

    _autosize(ws, {1: 52, 2: 62})


def submission_workbook(cfg: dict, link, submission, uploads_by_slot: dict) -> bytes:
    """One submission as the office's master file: Details, LLP agreement, ODI."""
    wb = Workbook()
    details = wb.active
    details.title = "Details"
    rows = _details_sheet(details, cfg, link, submission, uploads_by_slot)

    _llp_agreement_sheet(wb.create_sheet("LLP agreement"), cfg, rows)
    _odi_sheet(wb.create_sheet("ODI sheet"), cfg, rows)
    _llp_gen_sheet(wb.create_sheet("Details to be filled"), cfg, submission)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def all_submissions_workbook(cfg: dict, records: list[tuple]) -> bytes:
    """Every submission as one row, for bulk review. `records` is a list of
    (link, submission) pairs, newest first."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    fields = [f for s in cfg["sections"] for f in s["fields"]]
    header = ["Client / reference", "Submitted (UTC)"] + [
        f"{s['title']} — {f['label']}" for s in cfg["sections"] for f in s["fields"]
    ]
    _write_row(ws, 1, header, _HEAD, _HEAD_FILL)
    ws.freeze_panes = "C2"

    for row, (link, submission) in enumerate(records, start=2):
        data = submission.form_data or {}
        values = [link.label or "—", submission.submitted_at.strftime("%Y-%m-%d %H:%M")]
        values += [data.get(f["name"], "") for f in fields]
        _write_row(ws, row, values, _BODY)

    _autosize(ws, {1: 26, 2: 18, **{i: 30 for i in range(3, len(header) + 1)}})
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def attachments_zip(cfg: dict, link, submission) -> bytes:
    """Every uploaded document for one submission, decrypted, in a single ZIP.

    Names are prefixed with the slot they were uploaded against, so two files
    called scan.jpg from different partners stay distinguishable.
    """
    labels = {up["name"]: up["label"] for s in cfg["sections"] for up in s["uploads"]}
    folder = sanitize_filename(link.label or "submission") or "submission"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest = [
            f"Client / reference : {link.label or '—'}",
            f"Submitted (UTC)    : {submission.submitted_at.strftime('%Y-%m-%d %H:%M')}",
            f"Exported (UTC)     : {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
            "",
            "Files:",
        ]
        seen: set[str] = set()
        for upload in submission.uploads:
            label = labels.get(upload.field_key, upload.field_key)
            name = f"{sanitize_filename(label)}__{sanitize_filename(upload.original_filename)}"
            # a duplicate name would silently overwrite inside the archive
            base, n = name, 2
            while name in seen:
                name, n = f"{n}_{base}", n + 1
            seen.add(name)
            try:
                zf.writestr(f"{folder}/{name}", load_decrypted(upload.stored_filename))
                manifest.append(f"  {name}")
            except FileNotFoundError:
                # A missing blob must not sink the whole download.
                manifest.append(f"  {name}  [MISSING ON DISK]")
        zf.writestr(f"{folder}/manifest.txt", "\n".join(manifest) + "\n")

    return buf.getvalue()
