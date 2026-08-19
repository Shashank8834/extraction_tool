"""End-to-end tests for the LLP intake tool.

Runs without Docker/Postgres/Tesseract: SQLite + a temp upload dir + a stubbed
OCR extractor. Drives the real ASGI app over HTTP.

Run:  python backend/tests/test_e2e.py
Exits non-zero if any check fails.
"""
import asyncio
import io
import os
import sys
import tempfile

BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND)

tmp = tempfile.mkdtemp(prefix="llp_test_")
os.environ["DATABASE_URL"] = f"sqlite:///{os.path.join(tmp, 'test.db')}"
os.environ["SECRET_KEY"] = "test-secret"
os.environ["ADMIN_USERNAME"] = "admin"
os.environ["ADMIN_PASSWORD"] = "secret123"
os.environ["BASE_URL"] = "http://testserver"
os.environ["MAX_UPLOAD_MB"] = "15"
from cryptography.fernet import Fernet  # noqa: E402

os.environ["FILE_ENCRYPTION_KEY"] = Fernet.generate_key().decode()

import app.config as cfgmod  # noqa: E402

cfgmod.settings.upload_dir = os.path.join(tmp, "uploads")
import app.storage as storage  # noqa: E402

storage.settings.upload_dir = cfgmod.settings.upload_dir

# --- unit: parsers -----------------------------------------------------------
from app.ocr.parsers import (  # noqa: E402
    parse_aadhaar, parse_bank_statement, parse_pan, parse_utility_bill, verhoeff_valid,
)

PASS, FAIL = 0, 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS: {name}")
    else:
        FAIL += 1
        print(f"  FAIL: {name}")


print("== parsers ==")
pan = parse_pan("Name\nRAHUL KUMAR SHARMA\nFather's Name\nSURESH SHARMA\n15/08/1990\nABCDE1234F")
check("pan number", pan.get("pan_number") == "ABCDE1234F")
check("pan first/last split", pan.get("first_name") == "RAHUL" and pan.get("last_name") == "KUMAR SHARMA")
check("pan father split", pan.get("father_first_name") == "SURESH")
check("pan dob", pan.get("dob") == "15/08/1990")
# Current e-PAN cards print no "Name"/"Father's Name" labels at all — just the
# values in order. The label-only parser returned nothing on these, which is
# what "Couldn't auto-read this one" meant to the client.
epan = parse_pan("""INCOME TAX DEPARTMENT          GOVT. OF INDIA
Permanent Account Number Card
ABCDE1234F
ARPITA BHATT
MANOJ KUMAR BHATT
01/01/1990
Signature""")
check("e-PAN without labels: number", epan.get("pan_number") == "ABCDE1234F")
check("e-PAN without labels: name", epan.get("first_name") == "ARPITA" and epan.get("last_name") == "BHATT")
check("e-PAN without labels: father", epan.get("father_first_name") == "MANOJ")
check("e-PAN without labels: dob", epan.get("dob") == "01/01/1990")
# OCR often spaces the characters out; stripping spaces glues the number to the
# preceding word, so a word-boundary match finds nothing
spaced = parse_pan("""Permanent Account Number
A B C D E 1 2 3 4 F
PRIYA SHARMA
RAMESH SHARMA
02/03/1988""")
check("spaced-out PAN number read", spaced.get("pan_number") == "ABCDE1234F")
check("spaced-out PAN name read", spaced.get("first_name") == "PRIYA")

# OCR splits the two-column card header freely. Phrase matching let "INCOME TAX"
# through as the holder's name, which is worse than reading nothing: it fills the
# form confidently and marks it as read-from-document.
noisy = parse_pan("""INCOME TAX
DEPARTMENT
GOVT. OF
INDIA
ABCDE1234F
ARPITA BHATT
MANOJ BHATT
01/01/1990""")
check("split card header is not read as a name", noisy.get("first_name") == "ARPITA")
check("split card header is not read as a father", noisy.get("father_first_name") == "MANOJ")

aad = parse_aadhaar("Address:\n12 MG Road, Banjara Hills,\nHyderabad, Telangana - 500034\n2345 6789 0123")
check("aadhaar present_address", "500034" in (aad.get("present_address") or ""))
bank = parse_bank_statement("Address: 45 Jubilee Hills,\nHyderabad, Telangana 500033")
check("bank permanent_address", "500033" in (bank.get("permanent_address") or ""))
util = parse_utility_bill("Consumer Name: RAHUL SHARMA\nService Address: Plot 88, Gachibowli,\nHyderabad 500032")
check("utility office_address", "500032" in (util.get("office_address") or ""))
check("utility name", util.get("utility_name") == "RAHUL SHARMA")

# --- stub OCR for the web flow ----------------------------------------------
import app.routers.public as public  # noqa: E402


def fake_extract(extractor_name, data, content_type):
    canon = {
        "pan": {"first_name": "RAHUL", "last_name": "SHARMA", "father_first_name": "SURESH",
                "father_last_name": "SHARMA", "pan_number": "ABCDE1234F", "dob": "15/08/1990"},
        "aadhaar": {"present_address": "12 MG Road, Hyderabad 500034"},
        "bank_statement": {"permanent_address": "45 Jubilee Hills, Hyderabad 500033"},
        "utility_bill": {"office_address": "Plot 88, Gachibowli 500032", "utility_name": "RAHUL SHARMA"},
    }
    return ("raw ocr text", canon.get(extractor_name, {}))


public.extract = fake_extract

import httpx  # noqa: E402
from app.config_loader import get_form_config  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402
from app.models import SubmissionLink  # noqa: E402

loop = asyncio.new_event_loop()
asyncio.set_event_loop(loop)


def make_client():
    ac = httpx.AsyncClient(transport=httpx.ASGITransport(app=app), base_url="http://testserver")

    class Sync:
        def get(self, *a, **k):
            return loop.run_until_complete(ac.get(*a, **k))

        def post(self, *a, **k):
            return loop.run_until_complete(ac.post(*a, **k))

    return Sync()


client = make_client()
CFG = get_form_config()
db = SessionLocal()
JPG = (b"\xff\xd8\xff\xe0jpeg", "image/jpeg")


def sample_value(field):
    t = field.get("type")
    if field.get("numeric"):
        # lakh-style grouping, as a client would actually type it
        return "1,00,000"
    if t == "email":
        return "person@example.com"
    if t == "tel":
        # leading zero: the export must not turn this into a number
        return "09998887777"
    if field.get("key") == "pan_number":
        return "ABCDE1234F"
    if t == "select":
        # skip a leading "— Select —" placeholder: it is blank, so a required
        # select would fail validation
        return next(o for o in field["options"] if o)
    return "Sample " + field["key"]


def valid_data(skip=()):
    data = {}
    for section in CFG["sections"]:
        for field in section["fields"]:
            if field["name"] in skip:
                continue
            if field.get("required"):
                data[field["name"]] = sample_value(field)
    return data


def required_files():
    """Distinct bytes per slot: identical files across two partners are refused
    on purpose, so a fixture that reused one blob would be testing nothing."""
    files = {}
    for section in CFG["sections"]:
        for up in section["uploads"]:
            if up.get("required"):
                name = up["name"]
                files[name] = (up["key"] + ".jpg", JPG[0] + name.encode(), JPG[1])
    return files


def new_link(label):
    client.post("/admin/links", data={"label": label}, follow_redirects=True)
    db.expire_all()
    return db.query(SubmissionLink).filter(SubmissionLink.label == label).first()


print("== auth + link ==")
check("dashboard needs auth", client.get("/admin", follow_redirects=False).status_code == 303)
check("bad login 401", client.post("/admin/login", data={"username": "admin", "password": "x"},
                                    follow_redirects=False).status_code == 401)
client.post("/admin/login", data={"username": "admin", "password": "secret123"}, follow_redirects=False)
link = new_link("LLP Client")
token = link.token
check("link created", token is not None)

print("== form renders ==")
body = client.get(f"/f/{token}").text
for needle in ["Proposed LLP", "First Designated Partner", "Second Designated Partner",
               "Registered Office", 'data-slot="partner1__pan"', 'data-extract-url="/f/']:
    check(f"form contains {needle}", needle in body)

print("== live extract ==")
r = client.post(f"/f/{token}/extract", data={"slot": "partner1__pan"},
                files={"file": ("pan.jpg", *JPG)})
j = r.json()
check("extract fills first_name", j["fields"].get("partner1__first_name") == "RAHUL")
check("extract no cross-partner leak", not any(k.startswith("partner2__") for k in j["fields"]))
check("unknown slot 400", client.post(f"/f/{token}/extract", data={"slot": "no__pe"},
                                       files={"file": ("x.jpg", *JPG)}).status_code == 400)
check("bad type 415", client.post(f"/f/{token}/extract", data={"slot": "partner1__pan"},
                                   files={"file": ("x.html", b"<h1>", "text/html")}).status_code == 415)

print("== full submission ==")
r = client.post(f"/f/{token}", data=valid_data(), files=required_files())
check("submit ok", r.status_code == 200 and "Thank you" in r.text)
db.expire_all()
sub = db.query(SubmissionLink).filter(SubmissionLink.token == token).first().submission
check("persisted", sub is not None and sub.form_data.get("partner1__first_name"))
check("every required upload stored", len(sub.uploads) == len(required_files()))
check("upload extracted", {u.field_key: u for u in sub.uploads}["partner1__pan"].extracted_data.get("pan_number") == "ABCDE1234F")
check("relink locked (410)", client.get(f"/f/{token}").status_code == 410)

print("== validation ==")
l2 = new_link("L2")
check("missing field 400", client.post(f"/f/{l2.token}", data=valid_data(skip=("partner1__first_name",)),
                                        files=required_files()).status_code == 400)
l3 = new_link("L3")
d = valid_data(skip=("partner1__permanent_address", "partner1__place_of_birth"))
d["partner1__din"] = "DIN123"
check("DIN makes fields optional", client.post(f"/f/{l3.token}", data=d, files=required_files()).status_code == 200)
l4 = new_link("L4")
check("no DIN keeps them required", client.post(f"/f/{l4.token}",
      data=valid_data(skip=("partner1__permanent_address",)), files=required_files()).status_code == 400)

print("== data quality ==")
l5 = new_link("L5")
d = valid_data()
d["partner1__mobile"] = "na"
check("placeholder mobile rejected",
      client.post(f"/f/{l5.token}", data=d, files=required_files()).status_code == 400)
l6 = new_link("L6")
d = valid_data()
d["partner1__email"] = "nil"
check("placeholder email rejected",
      client.post(f"/f/{l6.token}", data=d, files=required_files()).status_code == 400)
l7 = new_link("L7")
d = valid_data()
d["partner1__mobile"] = "12345"
check("short mobile rejected",
      client.post(f"/f/{l7.token}", data=d, files=required_files()).status_code == 400)
l8 = new_link("L8")
d = valid_data()
d["partner1__pan_number"] = "1234567890"
check("malformed PAN rejected",
      client.post(f"/f/{l8.token}", data=d, files=required_files()).status_code == 400)
l13 = new_link("L13")
d13 = valid_data()
d13["partner1__pan_number"] = "abcde1234f"
r13 = client.post(f"/f/{l13.token}", data=d13, files=required_files())
db.expire_all()
sub13 = db.query(SubmissionLink).filter(SubmissionLink.token == l13.token).first().submission
check("PAN stored upper case", r13.status_code == 200
      and sub13.form_data.get("partner1__pan_number") == "ABCDE1234F")

l9 = new_link("L9")
d = valid_data()
d["partner1__mobile"] = "+91 99988 87777"
check("spaced +91 mobile accepted",
      client.post(f"/f/{l9.token}", data=d, files=required_files()).status_code == 200)
# one partner's document uploaded as another's: the failure that quietly copies
# the wrong address into a filing
l10 = new_link("L10")
files = required_files()
files["partner2__aadhaar"] = files["partner1__aadhaar"]
check("same identity document across two partners rejected",
      client.post(f"/f/{l10.token}", data=valid_data(), files=files).status_code == 400)
# spouses with a joint account file one statement; the office bill belongs to
# no partner at all. Refusing these blocked honest submissions.
l12 = new_link("L12")
files = required_files()
files["partner2__bank_statement"] = files["partner1__bank_statement"]
check("shared bank statement allowed",
      client.post(f"/f/{l12.token}", data=valid_data(), files=files).status_code == 200)
l11 = new_link("L11")
files = required_files()
files["partner1__bank_statement"] = files["partner1__aadhaar"]
check("same file within one partner allowed",
      client.post(f"/f/{l11.token}", data=valid_data(), files=files).status_code == 200)

print("== documents survive a failed submit ==")
from app.models import StagedUpload  # noqa: E402

PDF = (b"%PDF-1.4 fake", "application/pdf")


def staged_count(link):
    db.expire_all()
    return db.query(StagedUpload).filter(StagedUpload.link_id == link.id).count()


# The reported bug: one missing field, and every attached document was gone.
l18 = new_link("L18")
r18a = client.post(f"/f/{l18.token}", data=valid_data(skip=("partner1__first_name",)),
                   files=required_files())
check("incomplete submit still 400", r18a.status_code == 400)
check("documents held after the error", staged_count(l18) == len(required_files()))
check("the form says they are attached", "is attached" in r18a.text and "pan.jpg" in r18a.text)
# ...and the second attempt sends no files at all, as a browser does.
r18b = client.post(f"/f/{l18.token}", data=valid_data())
check("re-submit needs no re-attaching", r18b.status_code == 200 and "Thank you" in r18b.text)
db.expire_all()
sub18 = db.query(SubmissionLink).filter(SubmissionLink.token == l18.token).first().submission
check("held documents became the uploads", len(sub18.uploads) == len(required_files()))
check("held document keeps its filename",
      {u.field_key: u for u in sub18.uploads}["partner1__pan"].original_filename == "pan.jpg")
check("staging cleared after submit", staged_count(l18) == 0)

# A document chosen for autofill is held even if the client never gets as far
# as pressing Submit with it attached.
l19 = new_link("L19")
client.post(f"/f/{l19.token}/extract", data={"slot": "partner1__pan"},
            files={"file": ("autofill-pan.jpg", *JPG)})
check("autofill upload is held", staged_count(l19) == 1)
files19 = {k: v for k, v in required_files().items() if k != "partner1__pan"}
check("submit completes with the held document",
      client.post(f"/f/{l19.token}", data=valid_data(), files=files19).status_code == 200)

# A fresh attachment replaces the held one rather than being ignored.
l20 = new_link("L20")
client.post(f"/f/{l20.token}/extract", data={"slot": "partner1__pan"},
            files={"file": ("old.jpg", b"\xff\xd8\xff\xe0old", "image/jpeg")})
files20 = required_files()
files20["partner1__pan"] = ("new.jpg", b"\xff\xd8\xff\xe0new", "image/jpeg")
client.post(f"/f/{l20.token}", data=valid_data(), files=files20)
db.expire_all()
sub20 = db.query(SubmissionLink).filter(SubmissionLink.token == l20.token).first().submission
check("a re-attached document replaces the held one",
      {u.field_key: u for u in sub20.uploads}["partner1__pan"].original_filename == "new.jpg")

# Removing one: the only way back once a document is held server-side.
l21 = new_link("L21")
client.post(f"/f/{l21.token}/extract", data={"slot": "partner1__pan"},
            files={"file": ("wrong.jpg", *JPG)})
check("unstage ok", client.post(f"/f/{l21.token}/unstage",
                                data={"slot": "partner1__pan"}).status_code == 200)
check("removed document is gone", staged_count(l21) == 0)
check("and it is required again",
      client.post(f"/f/{l21.token}", data=valid_data(),
                  files={k: v for k, v in required_files().items()
                         if k != "partner1__pan"}).status_code == 400)

# A refused document is not held, or the client would see it as attached and
# hit the same refusal on every further attempt.
l23 = new_link("L23")
files23 = required_files()
files23["partner2__aadhaar"] = files23["partner1__aadhaar"]
r23 = client.post(f"/f/{l23.token}", data=valid_data(), files=files23)
check("shared identity document still refused", r23.status_code == 400)
db.expire_all()
held23 = {s.field_key for s in db.query(StagedUpload).filter(StagedUpload.link_id == l23.id)}
check("the refused document is not held", "partner2__aadhaar" not in held23)
check("the documents that were fine are held", "partner1__aadhaar" in held23)

print("== file types are read from the bytes ==")
from app.security import sniff_content_type  # noqa: E402

check("pdf sniffed", sniff_content_type(b"%PDF-1.7 ...", "application/octet-stream") == "application/pdf")
check("jpeg sniffed", sniff_content_type(b"\xff\xd8\xff\xe0\x00", "") == "image/jpeg")
check("png sniffed", sniff_content_type(b"\x89PNG\r\n\x1a\n", None) == "image/png")
check("heic sniffed", sniff_content_type(b"\x00\x00\x00\x18ftypheic", "application/octet-stream")
      == "image/heic")
check("declared type kept when bytes are unfamiliar",
      sniff_content_type(b"\x00\x01\x02\x03", "image/tiff") == "image/tiff")
check("a docx is named for what it is",
      sniff_content_type(b"PK\x03\x04zzz", "application/octet-stream") == "application/zip")

# The phone that sends a PDF as application/octet-stream used to be told its
# document was the wrong type; the bytes say otherwise.
l22 = new_link("L22")
check("octet-stream PDF accepted",
      client.post(f"/f/{l22.token}/extract", data={"slot": "partner1__pan"},
                  files={"file": ("scan.pdf", PDF[0], "application/octet-stream")}
                  ).status_code == 200)
check("a Word file is still refused",
      client.post(f"/f/{l22.token}/extract", data={"slot": "partner1__aadhaar"},
                  files={"file": ("cv.docx", b"PK\x03\x04zzz", "application/octet-stream")}
                  ).status_code == 415)
check("an image mislabelled as html is accepted on its bytes",
      client.post(f"/f/{l22.token}/extract", data={"slot": "partner1__aadhaar"},
                  files={"file": ("scan", JPG[0], "text/html")}).status_code == 200)

print("== master workbook ==")
from openpyxl import load_workbook  # noqa: E402
from app.exporting import submission_workbook  # noqa: E402

sub_link = db.query(SubmissionLink).filter(SubmissionLink.token == token).first()
xlsx = submission_workbook(CFG, sub_link, sub_link.submission,
                           {u.field_key: u for u in sub_link.submission.uploads})
wbk = load_workbook(io.BytesIO(xlsx))
check("five sheets", wbk.sheetnames == ["Details", "LLP agreement", "ODI sheet",
                                        "Details to be filled", "ODI to be filled"])
# llp-gen reads this sheet with plain openpyxl: a formula would arrive as its
# literal text and land in a filed document
gen = wbk["Details to be filled"]
check("bridge sheet holds values, not formulas",
      not any(isinstance(c.value, str) and c.value.startswith("=")
              for r in gen.iter_rows() for c in r))
check("bridge sheet keeps llp-gen row positions",
      gen["A1"].value == "Name of the Proposed LLP"
      and "designated partner details" in gen["A2"].value.lower()
      and "designated partner details" in gen["A19"].value.lower()
      and gen["A37"].value == "First Designated partner share"
      and gen["A44"].value == "Registered office address"
      and gen["A51"].value == "Objects of the LLP")
check("relation reaches the generator", gen["B18"].value in ("Son", "Daughter", "Wife", "Husband"))

# the ODI tab of the generator reads this one: Sl.No | Section | Field | Value
odi = wbk["ODI to be filled"]
check("ODI sheet uses the checklist header", odi["A1"].value == "Sl.No" and odi["C1"].value.startswith("Field Name"))
check("ODI sheet holds values, not formulas",
      not any(isinstance(c.value, str) and c.value.startswith("=")
              for r in odi.iter_rows() for c in r))
odi_rows = {odi.cell(row=r, column=3).value: odi.cell(row=r, column=4).value
            for r in range(2, odi.max_row + 1)}
check("ODI sheet carries the LLP name", bool(odi_rows.get("LLP Name")))
check("ODI sheet leaves the foreign entity to the team",
      not odi_rows.get("Foreign Entity Name"))
check("no Notes column", wbk["Details"]["C1"].value is None)

det = wbk["Details"]
labels = {det.cell(row=r, column=1).value: r for r in range(1, det.max_row + 1)}
cap_row = labels["First Partner — Capital Contribution"]
# a text cell counts as 0 inside SUM, so the capital total and every
# percentage on the agreement sheet would come out wrong
check("capital written as a number", isinstance(det.cell(row=cap_row, column=2).value, (int, float)))
phone_row = labels["LLP Contact Number (different from the ones above)"]
check("leading zero kept on phone", str(det.cell(row=phone_row, column=2).value).startswith("0"))

agr = wbk["LLP agreement"]
formulas = [c.value for r in agr.iter_rows() for c in r if isinstance(c.value, str) and c.value.startswith("=")]
check("agreement pulls from Details", any("Details!B" in f for f in formulas))
check("agreement totals the capital", any(f.startswith("=SUM(") for f in formulas))
check("blank source renders blank, not 0",
      all('=IF(Details!B' in f for f in formulas if "Details!B" in f))
# every reference must land on a row that exists, or the sheet shows #REF!
import re as _re
refs = [int(m) for f in formulas for m in _re.findall(r"Details!B(\d+)", f)]
check("no reference past the end of Details", refs and max(refs) <= det.max_row)
check("ODI keeps the team's blanks empty",
      all(wbk["ODI sheet"].cell(row=r, column=2).value is None
          for r in range(1, wbk["ODI sheet"].max_row + 1)
          if wbk["ODI sheet"].cell(row=r, column=1).value in ("SWIFT", "IBAN", "Foreign entity name")))

print("== security headers ==")
r = client.get("/admin")
check("CSP", "content-security-policy" in {k.lower() for k in r.headers})
check("X-Frame-Options", r.headers.get("x-frame-options") == "DENY")
check("nosniff", r.headers.get("x-content-type-options") == "nosniff")

print(f"\n==== {PASS} passed, {FAIL} failed ====")
sys.exit(1 if FAIL else 0)
